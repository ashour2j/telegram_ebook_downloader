#!/usr/bin/env python3
"""
Build the Excel book database + a JSON feed for the invoice website.

Scans the downloaded library, auto-counts the pages of every PDF, and writes:
  - books_database.xlsx  : Excel workbook (Books sheet + Settings sheet).
      * Pages ............ auto-counted from the PDF.
      * Paper Count ...... = ROUND(Pages * pages_per_paper, 0)   (linked to Settings)
      * Unit Price ....... = ROUND(Paper Count * price_per_paper
                                   + IF(Cover="Y", cover_cost, 0), 2)
      * Cover ............. Y/N per book - "N" removes the cover cost (booklet style).
  - books_data.json      : plain data the invoice website loads.
  - books.csv            : a flat CSV copy for quick imports.

Companion books (ملحق / مراجعة / حلول) are linked to their main book via a
"companion_of" field using (series, subject, grade, term, year).

Pricing (editable in the Settings sheet or via CLI flags):
    unit price = (paper_count * price_per_paper) + cover_cost (if cover = Y)
    paper_count = ROUND(pages * pages_per_paper, 0)
Defaults: price_per_paper = 0.65, cover_cost = 20, pages_per_paper = 0.5
(pages_per_paper = 0.5  ->  price = ((pages / 2) * 0.65) + 20  = double-sided printing)
"""

import argparse
import csv
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import DownloaderConfig
from rename_library import ALLOWED_EXTS, GRADE_ARABIC, TERM_ARABIC, build_target_name, parse_book

DEFAULT_RATE = 0.65
DEFAULT_COVER = 20.0
DEFAULT_PAGES_PER_PAPER = 0.5  # 0.5 = double-sided printing -> price = ((pages/2)*rate)+cover
DEFAULT_COVER_FLAG = "Y"

COMPANION_TYPES = ("ملحق", "مراجعة", "حلول")


def count_pdf_pages(path: Path) -> int:
    """Returns the number of pages in a PDF, or None if it can't be read."""
    try:
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        if reader.is_encrypted:
            try:
                reader.decrypt("")
            except Exception:
                return None
        return len(reader.pages)
    except Exception:
        return None


def _comp_key(book: dict):
    """Groups a book with its companions: same series/subject/grade/term/year."""
    return (
        book.get("series"),
        book.get("subject"),
        book.get("grade"),
        book.get("term"),
        book.get("year"),
    )


def collect_books(base_dir: Path):
    """Walks the library and returns (books, dropped_files).

    dropped_files lists the duplicate copies (same book downloaded more than
    once from different chats) that were collapsed into one listing.
    """
    files = sorted(
        p for p in base_dir.rglob("*")
        if p.is_file() and p.suffix.lower() in ALLOWED_EXTS
    )
    if not files:
        print(f"[!] No ebook files found under {base_dir}")
        return [], []

    raw = []
    for path in files:
        meta = parse_book(path.stem, path.parent)
        pages = count_pdf_pages(path) if path.suffix.lower() == ".pdf" else None
        raw.append({
            "series": meta["series"],
            "subject": meta["subject"],
            "grade": meta["grade"],
            "term": meta["term"],
            "year": meta["year"],
            "publisher": meta["publisher"],
            "type": meta["type"],
            "part": meta["part"],
            "extra": meta["extra"],
            "title": build_target_name(meta, path.suffix.lstrip(".")),
            "file": str(path.relative_to(base_dir)),
            "size_bytes": path.stat().st_size,
            "format": path.suffix.lower().lstrip("."),
            "pages": pages,
            "grade_arabic": GRADE_ARABIC.get(meta["grade"], ""),
            "term_arabic": TERM_ARABIC.get(meta["term"], ""),
        })

    # Drop exact duplicates (same book downloaded more than once from different chats).
    # "part" (the rename tool's "(2)" uniquifier) is ignored on purpose: it is a
    # filesystem marker, not a real volume number. Plain copies (no "(2)") are
    # preferred so the storefront shows clean titles.
    raw.sort(key=lambda b: (b["part"] is not None, b["file"]))
    seen = set()
    books = []
    dropped = []
    for book in raw:
        key = (book["series"], book["subject"], book["grade"], book["term"],
               book["year"], book["publisher"], book["type"], book["extra"])
        if key in seen:
            dropped.append(book["file"])
            continue
        seen.add(key)
        books.append(book)

    for idx, book in enumerate(books, start=1):
        book["id"] = idx
    return books, dropped


def assign_companions(books):
    """Sets companion_of for ملحق/مراجعة/حلول books pointing at their main book."""
    mains = {}
    for b in books:
        if b.get("type") == "كتاب":
            mains.setdefault(_comp_key(b), b)
    for b in books:
        b["companion_of"] = None
        if b.get("type") in COMPANION_TYPES:
            main = mains.get(_comp_key(b))
            if main:
                b["companion_of"] = main["id"]
    return books


def finalize_prices(books, rate, cover, pages_per_paper, default_cover="Y"):
    """Computes paper_count and both cover/no-cover prices for every book."""
    for book in books:
        pages = book["pages"] or 0
        book["paper_count"] = int(round(pages * pages_per_paper))
        book["has_cover"] = (default_cover or "Y").upper() == "Y"
        book["unit_price"] = round(book["paper_count"] * rate + cover, 2)
        book["no_cover_price"] = round(book["paper_count"] * rate, 2)
    return books


def load_settings(out_db: Path):
    """Reads pricing settings from an existing workbook (returns defaults if absent)."""
    result = {
        "rate": DEFAULT_RATE,
        "cover": DEFAULT_COVER,
        "pages_per_paper": DEFAULT_PAGES_PER_PAPER,
        "default_cover": DEFAULT_COVER_FLAG,
    }
    if not Path(out_db).exists():
        return result
    try:
        wb = openpyxl.load_workbook(out_db, data_only=False)
    except Exception:
        return result
    if "Settings" in wb.sheetnames:
        for row in wb["Settings"].iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            name = str(row[0]).strip()
            if name == "price_per_paper" and row[1] is not None:
                result["rate"] = float(row[1])
            elif name == "cover_cost" and row[1] is not None:
                result["cover"] = float(row[1])
            elif name == "pages_per_paper" and row[1] is not None:
                result["pages_per_paper"] = float(row[1])
            elif name == "default_cover" and row[1] is not None:
                result["default_cover"] = str(row[1]).strip().upper()
    wb.close()
    return result


def preserve_cover_flags(books, out_db):
    """Keeps per-book Cover (Y/N) edits made in the Excel file across rebuilds."""
    if not Path(out_db).exists():
        return books
    try:
        wb = openpyxl.load_workbook(out_db, data_only=False)
    except Exception:
        return books
    if "Books" not in wb.sheetnames:
        wb.close()
        return books
    ws = wb["Books"]
    headers = [c.value for c in ws[1]]
    i_file = headers.index("File") if "File" in headers else None
    i_cover = headers.index("Cover") if "Cover" in headers else None
    by_file = {}
    if i_file is not None and i_cover is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if i_file < len(row) and i_cover < len(row):
                file_val = row[i_file]
                cover_val = row[i_cover]
                if file_val is not None and cover_val is not None:
                    by_file[str(file_val)] = str(cover_val).strip().upper() == "Y"
    wb.close()
    for b in books:
        if b["file"] in by_file:
            b["has_cover"] = by_file[b["file"]]
    return books


def write_excel(books, path, rate, cover, pages_per_paper, default_cover="Y"):
    """Writes the Excel workbook with a Books sheet (formulas) and Settings sheet."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Books"
    headers = [
        "Book ID", "Series", "Subject", "Grade", "Term", "Year", "Publisher", "Type",
        "Cover", "File", "Format", "Pages", "Paper Count", "Unit Price (EGP)",
    ]
    ws.append(headers)
    for col, width in enumerate(
        [8, 18, 18, 8, 8, 8, 14, 10, 8, 60, 8, 8, 12, 14], start=1
    ):
        ws.column_dimensions[get_column_letter(col)].width = width

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5B8F")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, book in enumerate(books, start=2):
        ws.append([
            book["id"],
            book["series"],
            book["subject"],
            book["grade"],
            book["term"],
            book["year"],
            book["publisher"],
            book["type"],
            "Y" if book["has_cover"] else "N",
            book["file"],
            book["format"],
            book["pages"] if book["pages"] else "",
            f"=ROUND(L{i}*Settings!$B$4,0)",   # Paper Count
            f'=ROUND(M{i}*Settings!$B$2+IF(I{i}="Y",Settings!$B$3,0),2)',  # Unit Price
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    sws = wb.create_sheet("Settings")
    sws.append(["Setting", "Value", "Note"])
    sws.append(["price_per_paper", rate, "EGP per printed paper"])
    sws.append(["cover_cost", cover, "EGP added to a book when Cover = Y"])
    sws.append(["pages_per_paper", pages_per_paper, "1 = single-sided, 0.5 = double-sided"])
    sws.append(["default_cover", default_cover, "Y = new books include a cover by default"])
    sws.append(["price_formula", "(PaperCount * price_per_paper) + IF(Cover='Y', cover_cost, 0)",
                "Unit Price column uses this"])
    sws.column_dimensions["A"].width = 20
    sws.column_dimensions["B"].width = 12
    sws.column_dimensions["C"].width = 48
    for cell in sws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5B8F")

    wb.save(path)
    print(f"Excel database written: {path}")


def write_json(books, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(books, f, ensure_ascii=False, indent=2)
    print(f"JSON feed written: {path}")


def write_csv(books, path):
    columns = ["id", "series", "subject", "grade", "term", "year", "publisher",
               "type", "title", "pages", "paper_count", "unit_price",
               "no_cover_price", "has_cover", "companion_of", "file"]
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for book in books:
            writer.writerow(book)
    print(f"CSV copy written: {path}")


def build(base_dir, out_db, out_json, out_csv, rate=None, cover=None,
          pages_per_paper=None, default_cover=None, delete_duplicates=False):
    """Builds the Excel DB, JSON feed and CSV from the library.

    Any None pricing value falls back to the settings saved in an existing
    workbook, then to the module defaults. Returns the book list ([] on error).
    """
    saved = load_settings(out_db)
    rate = rate if rate is not None else saved["rate"]
    cover = cover if cover is not None else saved["cover"]
    pages_per_paper = pages_per_paper if pages_per_paper is not None else saved["pages_per_paper"]
    default_cover = default_cover if default_cover is not None else saved["default_cover"]

    print(f"Scanning library: {base_dir}")
    books, dropped = collect_books(base_dir)
    if not books:
        print("[!] No ebooks found; nothing to build.", file=sys.stderr)
        return []
    if delete_duplicates and dropped:
        for rel in dropped:
            dup_path = Path(base_dir) / rel
            try:
                dup_path.unlink()
                print(f"  deleted duplicate: {dup_path.name}")
            except OSError as exc:
                print(f"  [warn] could not delete {dup_path.name}: {exc}")
    books = assign_companions(books)
    books = finalize_prices(books, rate, cover, pages_per_paper, default_cover)
    books = preserve_cover_flags(books, out_db)
    print(f"Books indexed: {len(books)}")

    write_excel(books, out_db, rate, cover, pages_per_paper, default_cover)
    write_json(books, out_json)
    write_csv(books, out_csv)
    return books


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dir", type=Path, default=None,
                        help="Library root to scan (default: OUTPUT_DIR from .env)")
    parser.add_argument("--out-db", type=Path, default=Path("books_database.xlsx"))
    parser.add_argument("--out-json", type=Path, default=Path("books_data.json"))
    parser.add_argument("--out-csv", type=Path, default=Path("books.csv"))
    parser.add_argument("--rate", type=float, default=None, help="EGP per paper")
    parser.add_argument("--cover", type=float, default=None, help="EGP cover cost per book")
    parser.add_argument("--pages-per-paper", type=float, default=None,
                        help="0.5 = double-sided printing, 1 = single-sided")
    parser.add_argument("--cover-default", choices=["Y", "N"], default=None,
                        help="Default cover for new books (Y/N)")
    parser.add_argument("--delete-duplicates", action="store_true",
                        help="Delete duplicate files (same book downloaded twice)")
    args = parser.parse_args()

    base_dir = args.dir if args.dir else DownloaderConfig().output_dir
    if not base_dir.is_dir():
        print(f"[!] Directory not found: {base_dir}", file=sys.stderr)
        sys.exit(1)

    books = build(
        base_dir, args.out_db, args.out_json, args.out_csv,
        rate=args.rate, cover=args.cover, pages_per_paper=args.pages_per_paper,
        default_cover=args.cover_default, delete_duplicates=args.delete_duplicates,
    )
    if not books:
        sys.exit(1)


if __name__ == "__main__":
    main()
